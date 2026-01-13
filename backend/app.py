from flask import Flask, jsonify, request, session, send_file
from flask_cors import CORS
import pymysql
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
import io
from datetime import datetime
import re

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'your-secret-key')  # 用于会话加密
app.config['JSON_AS_ASCII'] = False  # 确保JSON响应不转义中文
CORS(app, supports_credentials=True)  # 配置跨域访问，支持凭证

# 数据库连接配置
db_config = {
    'host': os.environ.get('DB_HOST', 'localhost'),
    'user': os.environ.get('DB_USER', 'root'),
    'password': os.environ.get('DB_PASSWORD', 'password'),
    'database': os.environ.get('DB_NAME', 'zhixinstudentsaas'),
    'charset': 'utf8mb4'
}

# 创建数据库连接
def get_db_connection():
    connection = pymysql.connect(**db_config)
    return connection

# API接口：用户登录
@app.route('/api/login', methods=['POST'])
def login():
    connection = None
    cursor = None
    try:
        # 获取请求参数
        data = request.get_json()
        username = data.get('username')
        password = data.get('password')
        
        if not username or not password:
            return jsonify({'error': '用户名和密码不能为空'}), 400
            
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)
        
        # 查询用户
        cursor.execute("SELECT username, status FROM useraccount WHERE username = %s AND password = %s", (username, password))
        user = cursor.fetchone()

        if user:
            # 检查账号状态
            if user['status'] == 1:
                return jsonify({'error': '该账号已被禁用'}), 403
            # 登录成功，保存到会话
            session['username'] = user['username']
            return jsonify({'message': '登录成功', 'username': user['username']}), 200
        else:
            return jsonify({'error': '用户名或密码错误'}), 401
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取当前登录用户信息
@app.route('/api/profile', methods=['GET'])
def get_profile():
    if 'username' in session:
        return jsonify({'username': session['username']}), 200
    else:
        return jsonify({'error': '未登录'}), 401

# API接口：用户登出
@app.route('/api/logout', methods=['POST'])
def logout():
    session.pop('username', None)
    return jsonify({'message': '登出成功'}), 200

# API接口：获取账号列表
@app.route('/api/accounts', methods=['GET'])
def get_accounts():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询所有账号
        cursor.execute("SELECT id, username, password, status FROM useraccount ORDER BY id")
        accounts = cursor.fetchall()

        return jsonify({'accounts': accounts}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新账号状态
@app.route('/api/accounts/<int:id>/status', methods=['PUT'])
def update_account_status(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status not in [0, 1]:
            return jsonify({'error': '状态值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新账号状态
        cursor.execute("UPDATE useraccount SET status = %s WHERE id = %s", (status, id))
        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '账号不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取学生列表
@app.route('/api/students', methods=['GET'])
def get_students():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)
        
        # 查询学生列表，包括关联的性别、年级和教练信息
        cursor.execute("""
            SELECT
                s.id,
                s.name AS student_name,
                s.sex_id,
                sx.name AS sex,
                s.grade_id,
                g.name AS grade,
                s.phone,
                s.status,
                GROUP_CONCAT(c.name SEPARATOR ', ') AS coach_names
            FROM
                student s
            JOIN
                sex sx ON s.sex_id = sx.id
            JOIN
                grade g ON s.grade_id = g.id
            LEFT JOIN
                student_coach sc ON s.id = sc.student_id
            LEFT JOIN
                coach c ON sc.coach_id = c.id
            GROUP BY
                s.id, s.name, s.sex_id, sx.name, s.grade_id, g.name, s.phone, s.status
        """)
        
        students = cursor.fetchall()
        return jsonify({'students': students}), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取性别列表
@app.route('/api/sexes', methods=['GET'])
def get_sexes():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        cursor.execute("SELECT id, name FROM sex ORDER BY id")
        sexes = cursor.fetchall()

        return jsonify({'sexes': sexes}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的年级列表
@app.route('/api/grades/active', methods=['GET'])
def get_active_grades():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询状态为启用(0)的年级
        cursor.execute("SELECT id, name AS grade FROM grade WHERE status = 0 ORDER BY id")
        grades = cursor.fetchall()

        return jsonify({'grades': grades}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的学科列表
@app.route('/api/subjects/active', methods=['GET'])
def get_active_subjects():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询状态为启用(0)的学科
        cursor.execute("SELECT id, subject FROM subject WHERE status = 0 ORDER BY id")
        subjects = cursor.fetchall()

        return jsonify({'subjects': subjects}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取教练列表
@app.route('/api/coaches', methods=['GET'])
def get_coaches():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询教练列表，包括关联的性别和学科信息
        cursor.execute("""
            SELECT
                c.id,
                c.name AS coach_name,
                c.sex_id,
                sx.name AS sex,
                c.subject_id,
                sub.subject,
                c.phone,
                c.status
            FROM
                coach c
            JOIN
                sex sx ON c.sex_id = sx.id
            JOIN
                subject sub ON c.subject_id = sub.id
        """)

        coaches = cursor.fetchall()
        return jsonify({'coaches': coaches}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新学生信息
@app.route('/api/students/<int:id>', methods=['PUT'])
def update_student(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        student_name = data.get('student_name')
        sex_id = data.get('sex_id')
        phone = data.get('phone')
        grade_id = data.get('grade_id')

        if not student_name or sex_id is None or not phone or grade_id is None:
            return jsonify({'error': '所有字段不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新学生信息
        cursor.execute("""
            UPDATE student
            SET name = %s, sex_id = %s, phone = %s, grade_id = %s
            WHERE id = %s
        """, (student_name, sex_id, phone, grade_id, id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '学生不存在'}), 404

        return jsonify({'message': '学生信息更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新学生状态
@app.route('/api/students/<int:id>/status', methods=['PUT'])
def update_student_status(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status not in [0, 1]:
            return jsonify({'error': '状态值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新学生状态
        cursor.execute("UPDATE student SET status = %s WHERE id = %s", (status, id))
        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '学生不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：删除学生
@app.route('/api/students/<int:id>', methods=['DELETE'])
def delete_student(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查是否有关联订单
        cursor.execute("SELECT COUNT(*) as count FROM orders WHERE student_id = %s", (id,))
        result = cursor.fetchone()
        if result['count'] > 0:
            return jsonify({'error': '该学生有关联订单，无法删除'}), 400

        # 删除学生与教练的关联
        cursor.execute("DELETE FROM student_coach WHERE student_id = %s", (id,))

        # 删除学生
        cursor.execute("DELETE FROM student WHERE id = %s", (id,))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '学生不存在'}), 404

        return jsonify({'message': '学生删除成功'}), 200
        
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新教练信息
@app.route('/api/coaches/<int:id>', methods=['PUT'])
def update_coach(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        coach_name = data.get('coach_name')
        sex_id = data.get('sex_id')
        phone = data.get('phone')
        subject_id = data.get('subject_id')

        if not coach_name or sex_id is None or not phone or subject_id is None:
            return jsonify({'error': '所有字段不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新教练信息
        cursor.execute("""
            UPDATE coach
            SET name = %s, sex_id = %s, phone = %s, subject_id = %s
            WHERE id = %s
        """, (coach_name, sex_id, phone, subject_id, id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '教练不存在'}), 404

        return jsonify({'message': '教练信息更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新教练状态
@app.route('/api/coaches/<int:id>/status', methods=['PUT'])
def update_coach_status(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status not in [0, 1]:
            return jsonify({'error': '状态值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新教练状态
        cursor.execute("UPDATE coach SET status = %s WHERE id = %s", (status, id))
        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '教练不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：删除教练
@app.route('/api/coaches/<int:id>', methods=['DELETE'])
def delete_coach(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # 删除教练与学生的关联
        cursor.execute("DELETE FROM student_coach WHERE coach_id = %s", (id,))

        # 删除教练
        cursor.execute("DELETE FROM coach WHERE id = %s", (id,))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '教练不存在'}), 404

        return jsonify({'message': '教练删除成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用状态的教练列表
@app.route('/api/coaches/active', methods=['GET'])
def get_active_coaches():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询启用状态的教练列表
        cursor.execute("""
            SELECT
                c.id,
                c.name AS coach_name
            FROM
                coach c
            WHERE
                c.status = '启用'
        """)

        coaches = cursor.fetchall()
        return jsonify({'coaches': coaches}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用状态的学生列表
@app.route('/api/students/active', methods=['GET'])
def get_active_students():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询启用状态的学生列表
        cursor.execute("""
            SELECT
                s.id,
                s.name AS student_name
            FROM
                student s
            WHERE
                s.status = '启用'
        """)

        students = cursor.fetchall()
        return jsonify({'students': students}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增教练
@app.route('/api/coaches', methods=['POST'])
def add_coach():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        coach_name = data.get('coach_name')
        sex_id = data.get('sex_id')
        phone = data.get('phone')
        subject_id = data.get('subject_id')
        student_ids = data.get('student_ids', [])  # 学生ID列表，非必填

        # 验证必填字段
        if not coach_name or sex_id is None or not phone or subject_id is None:
            return jsonify({'error': '姓名、性别、电话和学科不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 插入教练信息
        cursor.execute("""
            INSERT INTO coach (name, sex_id, subject_id, phone, status)
            VALUES (%s, %s, %s, %s, 0)
        """, (coach_name, sex_id, subject_id, phone))

        coach_id = cursor.lastrowid

        # 如果有选择学生，插入学生与教练的关联
        if student_ids:
            for student_id in student_ids:
                cursor.execute("""
                    INSERT INTO student_coach (student_id, coach_id)
                    VALUES (%s, %s)
                """, (student_id, coach_id))

        connection.commit()

        return jsonify({'message': '教练添加成功', 'coach_id': coach_id}), 201

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增学生
@app.route('/api/students', methods=['POST'])
def add_student():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        student_name = data.get('student_name')
        sex_id = data.get('sex_id')
        phone = data.get('phone')
        grade_id = data.get('grade_id')
        coach_ids = data.get('coach_ids', [])  # 教练ID列表，非必填

        # 验证必填字段
        if not student_name or sex_id is None or not phone or grade_id is None:
            return jsonify({'error': '姓名、性别、电话和年级不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 插入学生信息
        cursor.execute("""
            INSERT INTO student (name, sex_id, grade_id, phone)
            VALUES (%s, %s, %s, %s)
        """, (student_name, sex_id, grade_id, phone))

        student_id = cursor.lastrowid

        # 如果有选择教练，插入学生与教练的关联
        if coach_ids:
            for coach_id in coach_ids:
                cursor.execute("""
                    INSERT INTO student_coach (student_id, coach_id)
                    VALUES (%s, %s)
                """, (student_id, coach_id))

        connection.commit()

        return jsonify({'message': '学生添加成功', 'student_id': student_id}), 201

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 订单管理API ====================

# API接口：获取订单列表
@app.route('/api/orders', methods=['GET'])
def get_orders():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询订单列表，关联学生信息（包含预计付款时间和优惠金额）
        cursor.execute("""
            SELECT
                o.id,
                o.student_id AS uid,
                s.name AS student_name,
                o.expected_payment_time,
                o.amount_receivable,
                o.discount_amount,
                o.amount_received,
                o.create_time,
                o.status
            FROM
                `orders` o
            JOIN
                student s ON o.student_id = s.id
            ORDER BY
                o.create_time DESC
        """)

        orders = cursor.fetchall()
        return jsonify({'orders': orders}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增订单（改造版：支持商品列表）
@app.route('/api/orders', methods=['POST'])
def create_order():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        goods_list = data.get('goods_list', [])  # 商品列表 [{goods_id, total_price, price}]
        expected_payment_time = data.get('expected_payment_time')  # 新增：预计付款时间
        activity_ids = data.get('activity_ids', [])  # 新增：活动ID列表
        discount_amount = float(data.get('discount_amount', 0))  # 新增：优惠金额
        child_discounts = data.get('child_discounts', {})  # 新增：子订单优惠分摊

        if not student_id:
            return jsonify({'error': '学生ID不能为空'}), 400

        if not goods_list or len(goods_list) == 0:
            return jsonify({'error': '必须至少选择一个商品'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 计算应收金额（所有商品的商品总价之和）
        total_receivable = sum(float(g.get('total_price', 0)) for g in goods_list)
        # 计算实收金额（所有商品的标准售价之和 - 优惠金额）
        total_received = sum(float(g.get('price', 0)) for g in goods_list) - discount_amount

        # 插入订单数据，状态默认为10（草稿），包含预计付款时间和优惠金额
        cursor.execute("""
            INSERT INTO `orders` (student_id, expected_payment_time, amount_receivable, amount_received, discount_amount, status)
            VALUES (%s, %s, %s, %s, %s, 10)
        """, (student_id, expected_payment_time, total_receivable, total_received, discount_amount))

        order_id = cursor.lastrowid

        # 创建子产品订单（包含优惠金额）
        for goods in goods_list:
            goods_id = goods.get('goods_id')
            goods_total_price = float(goods.get('total_price', 0))
            goods_price = float(goods.get('price', 0))

            # 获取该商品的优惠金额
            child_discount = float(child_discounts.get(str(goods_id), 0))
            # 计算子订单实收金额
            child_received = goods_price - child_discount

            cursor.execute("""
                INSERT INTO childorders (parentsid, goodsid, amount_receivable, amount_received, discount_amount, status)
                VALUES (%s, %s, %s, %s, %s, 10)
            """, (order_id, goods_id, goods_total_price, child_received, child_discount))

        # 批量插入订单活动关联
        for activity_id in activity_ids:
            cursor.execute("""
                INSERT INTO orders_activity (orders_id, activity_id)
                VALUES (%s, %s)
            """, (order_id, activity_id))

        connection.commit()

        return jsonify({'message': '订单创建成功', 'order_id': order_id}), 201

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取订单的商品列表
@app.route('/api/orders/<int:order_id>/goods', methods=['GET'])
def get_order_goods(order_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询订单的子商品列表（包含优惠金额）
        cursor.execute("""
            SELECT
                c.id,
                c.goodsid,
                g.name AS goods_name,
                g.isgroup,
                b.name AS brand_name,
                cl.name AS classify_name,
                c.amount_receivable,
                c.discount_amount,
                c.amount_received
            FROM childorders c
            JOIN goods g ON c.goodsid = g.id
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify cl ON g.classifyid = cl.id
            WHERE c.parentsid = %s
            ORDER BY c.id
        """, (order_id,))

        goods_list = cursor.fetchall()

        # 为每个商品获取属性信息
        for goods in goods_list:
            cursor.execute("""
                SELECT
                    a.name AS attr_name,
                    av.name AS value_name
                FROM goods_attributevalue gav
                JOIN attribute_value av ON gav.attributevalueid = av.id
                JOIN attribute a ON av.attributeid = a.id
                WHERE gav.goodsid = %s AND a.classify = 0
            """, (goods['goodsid'],))
            attr_values = cursor.fetchall()
            attributes = [f"{av['attr_name']}:{av['value_name']}" for av in attr_values]
            goods['attributes'] = ', '.join(attributes) if attributes else ''

        return jsonify({'goods': goods_list}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新订单（支持商品列表更新）
@app.route('/api/orders/<int:order_id>', methods=['PUT'])
def update_order(order_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        goods_list = data.get('goods_list', [])
        expected_payment_time = data.get('expected_payment_time')  # 新增：预计付款时间
        activity_ids = data.get('activity_ids', [])  # 新增：活动ID列表
        discount_amount = float(data.get('discount_amount', 0))  # 新增：优惠金额
        child_discounts = data.get('child_discounts', {})  # 新增：子订单优惠分摊

        if not goods_list or len(goods_list) == 0:
            return jsonify({'error': '必须至少选择一个商品'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查订单状态是否为草稿
        cursor.execute("SELECT status FROM `orders` WHERE id = %s", (order_id,))
        order = cursor.fetchone()

        if not order:
            return jsonify({'error': '订单不存在'}), 404

        if order['status'] != 10:
            return jsonify({'error': '只能编辑草稿状态的订单'}), 400

        # 计算新的应收金额和实收金额
        total_receivable = sum(float(g.get('total_price', 0)) for g in goods_list)
        total_received = sum(float(g.get('price', 0)) for g in goods_list) - discount_amount

        # 更新订单金额、预计付款时间和优惠金额
        cursor.execute("""
            UPDATE `orders`
            SET amount_receivable = %s, amount_received = %s, discount_amount = %s, expected_payment_time = %s
            WHERE id = %s
        """, (total_receivable, total_received, discount_amount, expected_payment_time, order_id))

        # 删除原有的子订单
        cursor.execute("DELETE FROM childorders WHERE parentsid = %s", (order_id,))

        # 删除原有的订单活动关联
        cursor.execute("DELETE FROM orders_activity WHERE orders_id = %s", (order_id,))

        # 创建新的子订单（包含优惠金额）
        for goods in goods_list:
            goods_id = goods.get('goods_id')
            goods_total_price = float(goods.get('total_price', 0))
            goods_price = float(goods.get('price', 0))

            # 获取该商品的优惠金额
            child_discount = float(child_discounts.get(str(goods_id), 0))
            # 计算子订单实收金额
            child_received = goods_price - child_discount

            cursor.execute("""
                INSERT INTO childorders (parentsid, goodsid, amount_receivable, amount_received, discount_amount, status)
                VALUES (%s, %s, %s, %s, %s, 10)
            """, (order_id, goods_id, goods_total_price, child_received, child_discount))

        # 批量插入新的订单活动关联
        for activity_id in activity_ids:
            cursor.execute("""
                INSERT INTO orders_activity (orders_id, activity_id)
                VALUES (%s, %s)
            """, (order_id, activity_id))

        connection.commit()

        return jsonify({'message': '订单更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：作废订单
@app.route('/api/orders/<int:order_id>/cancel', methods=['PUT'])
def cancel_order(order_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查订单状态是否为草稿
        cursor.execute("SELECT status FROM `orders` WHERE id = %s", (order_id,))
        order = cursor.fetchone()

        if not order:
            return jsonify({'error': '订单不存在'}), 404

        if order['status'] != 10:
            return jsonify({'error': '只能作废草稿状态的订单'}), 400

        # 将订单状态更新为99（已作废）
        cursor.execute("""
            UPDATE `orders`
            SET status = 99
            WHERE id = %s
        """, (order_id,))

        # 同时作废关联的子产品订单
        cursor.execute("""
            UPDATE childorders
            SET status = 99
            WHERE parentsid = %s
        """, (order_id,))

        connection.commit()

        return jsonify({'message': '订单已作废'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：提交订单
@app.route('/api/orders/<int:order_id>/submit', methods=['PUT'])
def submit_order(order_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查订单状态是否为草稿
        cursor.execute("SELECT status FROM `orders` WHERE id = %s", (order_id,))
        order = cursor.fetchone()

        if not order:
            return jsonify({'error': '订单不存在'}), 404

        if order['status'] != 10:
            return jsonify({'error': '只能提交草稿状态的订单'}), 400

        # 将订单状态更新为20（未支付）
        cursor.execute("""
            UPDATE `orders`
            SET status = 20
            WHERE id = %s
        """, (order_id,))

        # 同时更新关联的子产品订单状态
        cursor.execute("""
            UPDATE childorders
            SET status = 20
            WHERE parentsid = %s
        """, (order_id,))

        connection.commit()

        return jsonify({'message': '订单已提交'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取子产品订单列表
@app.route('/api/childorders', methods=['GET'])
def get_childorders():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询子产品订单列表，关联商品信息（包含优惠金额）
        cursor.execute("""
            SELECT
                c.id,
                c.parentsid,
                c.goodsid,
                g.name AS goods_name,
                c.amount_receivable,
                c.discount_amount,
                c.amount_received,
                c.status,
                c.create_time
            FROM
                childorders c
            JOIN
                goods g ON c.goodsid = g.id
            ORDER BY
                c.id DESC
        """)

        childorders = cursor.fetchall()
        return jsonify({'childorders': childorders}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取商品总价（用于订单选择商品时计算）
@app.route('/api/goods/<int:goods_id>/total-price', methods=['GET'])
def get_goods_total_price(goods_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取商品基本信息
        cursor.execute("""
            SELECT id, name, isgroup, price
            FROM goods
            WHERE id = %s
        """, (goods_id,))

        goods = cursor.fetchone()
        if not goods:
            return jsonify({'error': '商品不存在'}), 404

        total_price = float(goods['price'])

        # 如果是组合商品，计算子商品标准售价之和
        if goods['isgroup'] == 0:
            cursor.execute("""
                SELECT SUM(g.price) AS total
                FROM goods_goods gg
                JOIN goods g ON gg.goodsid = g.id
                WHERE gg.parentsid = %s
            """, (goods_id,))
            result = cursor.fetchone()
            if result and result['total']:
                total_price = float(result['total'])

        return jsonify({
            'goods_id': goods_id,
            'price': float(goods['price']),
            'total_price': total_price,
            'isgroup': goods['isgroup']
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用状态的商品列表（用于订单选择商品）
@app.route('/api/goods/active-for-order', methods=['GET'])
def get_active_goods_for_order():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取所有启用状态的商品
        cursor.execute("""
            SELECT
                g.id, g.name, g.brandid, g.classifyid, g.isgroup, g.price,
                b.name AS brand_name,
                c.name AS classify_name
            FROM goods g
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify c ON g.classifyid = c.id
            WHERE g.status = 0
            ORDER BY g.id DESC
        """)

        goods_list = cursor.fetchall()

        # 为每个商品获取属性信息和总价
        for goods in goods_list:
            # 获取属性值
            cursor.execute("""
                SELECT
                    av.id, av.name AS value_name, av.attributeid,
                    a.name AS attr_name, a.classify
                FROM goods_attributevalue gav
                JOIN attribute_value av ON gav.attributevalueid = av.id
                JOIN attribute a ON av.attributeid = a.id
                WHERE gav.goodsid = %s
            """, (goods['id'],))

            attr_values = cursor.fetchall()
            attributes = []
            for av in attr_values:
                if av['classify'] == 0:  # 属性
                    attributes.append(f"{av['attr_name']}:{av['value_name']}")
            goods['attributes'] = ', '.join(attributes) if attributes else ''

            # 计算总价
            if goods['isgroup'] == 0:  # 组合商品
                cursor.execute("""
                    SELECT SUM(g.price) AS total
                    FROM goods_goods gg
                    JOIN goods g ON gg.goodsid = g.id
                    WHERE gg.parentsid = %s
                """, (goods['id'],))
                result = cursor.fetchone()
                goods['total_price'] = float(result['total']) if result and result['total'] else float(goods['price'])
            else:
                goods['total_price'] = float(goods['price'])

        return jsonify({'goods': goods_list}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 属性管理API ====================

# API接口：获取属性列表
@app.route('/api/attributes', methods=['GET'])
def get_attributes():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询所有属性，包含属性值数量
        cursor.execute("""
            SELECT a.id, a.name, a.classify, a.status, a.create_time, a.update_time,
                   COUNT(av.id) AS value_count
            FROM attribute a
            LEFT JOIN attribute_value av ON a.id = av.attributeid
            GROUP BY a.id, a.name, a.classify, a.status, a.create_time, a.update_time
            ORDER BY a.id DESC
        """)
        attributes = cursor.fetchall()

        return jsonify({'attributes': attributes}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增属性
@app.route('/api/attributes', methods=['POST'])
def create_attribute():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        classify = data.get('classify')

        if not name or classify is None:
            return jsonify({'error': '名称和分类不能为空'}), 400

        if classify not in [0, 1]:
            return jsonify({'error': '分类值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 插入属性数据，状态默认为0（启用）
        cursor.execute("""
            INSERT INTO attribute (name, classify, status)
            VALUES (%s, %s, 0)
        """, (name, classify))

        connection.commit()
        attribute_id = cursor.lastrowid

        return jsonify({'message': '属性创建成功', 'attribute_id': attribute_id}), 201

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新属性
@app.route('/api/attributes/<int:attribute_id>', methods=['PUT'])
def update_attribute(attribute_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        classify = data.get('classify')

        if not name or classify is None:
            return jsonify({'error': '名称和分类不能为空'}), 400

        if classify not in [0, 1]:
            return jsonify({'error': '分类值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新属性
        cursor.execute("""
            UPDATE attribute
            SET name = %s, classify = %s
            WHERE id = %s
        """, (name, classify, attribute_id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '属性不存在'}), 404

        return jsonify({'message': '属性更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新属性状态
@app.route('/api/attributes/<int:attribute_id>/status', methods=['PUT'])
def update_attribute_status(attribute_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status not in [0, 1]:
            return jsonify({'error': '状态值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新属性状态
        cursor.execute("""
            UPDATE attribute
            SET status = %s
            WHERE id = %s
        """, (status, attribute_id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '属性不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 属性值管理API ====================

# API接口：获取指定属性的属性值列表
@app.route('/api/attributes/<int:attribute_id>/values', methods=['GET'])
def get_attribute_values(attribute_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询该属性的所有属性值
        cursor.execute("""
            SELECT id, name, attributeid
            FROM attribute_value
            WHERE attributeid = %s
            ORDER BY id ASC
        """, (attribute_id,))
        values = cursor.fetchall()

        return jsonify({'values': values}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：保存属性值（批量更新）
@app.route('/api/attributes/<int:attribute_id>/values', methods=['POST'])
def save_attribute_values(attribute_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        values = data.get('values', [])

        # 验证至少有一个属性值
        if not values or len(values) == 0:
            return jsonify({'error': '至少需要填入一条属性值'}), 400

        # 验证所有属性值不为空
        for value in values:
            if not value or not value.strip():
                return jsonify({'error': '属性值不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 先删除该属性的所有旧属性值
        cursor.execute("DELETE FROM attribute_value WHERE attributeid = %s", (attribute_id,))

        # 插入新的属性值
        for value in values:
            cursor.execute("""
                INSERT INTO attribute_value (name, attributeid)
                VALUES (%s, %s)
            """, (value.strip(), attribute_id))

        connection.commit()

        return jsonify({'message': '属性值保存成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 类型管理API ====================

# API接口：获取类型列表
@app.route('/api/classifies', methods=['GET'])
def get_classifies():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询所有类型，关联父级名称
        cursor.execute("""
            SELECT c.id, c.name, c.level, c.parentid AS parent_id, c.status,
                   p.name AS parent_name
            FROM classify c
            LEFT JOIN classify p ON c.parentid = p.id
            ORDER BY c.level ASC, c.id DESC
        """)
        classifies = cursor.fetchall()

        return jsonify({'classifies': classifies}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的一级类型列表（用于新增/编辑二级类型时选择父级）
@app.route('/api/classifies/parents', methods=['GET'])
def get_parent_classifies():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询状态为启用(0)且级别为一级(0)的类型
        cursor.execute("""
            SELECT id, name
            FROM classify
            WHERE status = 0 AND level = 0
            ORDER BY id ASC
        """)
        parents = cursor.fetchall()

        return jsonify({'parents': parents}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增类型
@app.route('/api/classifies', methods=['POST'])
def create_classify():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        level = data.get('level')
        parent_id = data.get('parent_id')

        if not name or level is None:
            return jsonify({'error': '名称和级别不能为空'}), 400

        if level not in [0, 1]:
            return jsonify({'error': '级别值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 校验名称是否重复
        if level == 0:
            # 一级类型：检查一级类型中名称是否已存在
            cursor.execute("""
                SELECT id FROM classify WHERE name = %s AND level = 0
            """, (name,))
            existing = cursor.fetchone()
            if existing:
                return jsonify({'error': '该一级类型名称已存在'}), 400
            parent_id = None  # 一级类型没有父级
        else:
            # 二级类型：检查同一父级下是否有同名的二级类型
            if not parent_id:
                return jsonify({'error': '二级类型必须选择父级类型'}), 400
            cursor.execute("""
                SELECT id FROM classify WHERE name = %s AND level = 1 AND parentid = %s
            """, (name, parent_id))
            existing = cursor.fetchone()
            if existing:
                return jsonify({'error': '该父级类型下已存在同名的二级类型'}), 400

        # 插入类型数据
        cursor.execute("""
            INSERT INTO classify (name, level, parentid, status)
            VALUES (%s, %s, %s, 0)
        """, (name, level, parent_id))

        connection.commit()
        classify_id = cursor.lastrowid

        return jsonify({'message': '类型创建成功', 'classify_id': classify_id}), 201

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新类型
@app.route('/api/classifies/<int:classify_id>', methods=['PUT'])
def update_classify(classify_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        level = data.get('level')
        parent_id = data.get('parent_id')

        if not name or level is None:
            return jsonify({'error': '名称和级别不能为空'}), 400

        if level not in [0, 1]:
            return jsonify({'error': '级别值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查类型是否存在
        cursor.execute("SELECT id, level FROM classify WHERE id = %s", (classify_id,))
        existing = cursor.fetchone()
        if not existing:
            return jsonify({'error': '类型不存在'}), 404

        # 校验名称是否重复（排除自身）
        if level == 0:
            # 一级类型：检查一级类型中名称是否已存在
            cursor.execute("""
                SELECT id FROM classify WHERE name = %s AND level = 0 AND id != %s
            """, (name, classify_id))
            duplicate = cursor.fetchone()
            if duplicate:
                return jsonify({'error': '该一级类型名称已存在'}), 400
            parent_id = None  # 一级类型没有父级
        else:
            # 二级类型：检查同一父级下是否有同名的二级类型
            if not parent_id:
                return jsonify({'error': '二级类型必须选择父级类型'}), 400
            cursor.execute("""
                SELECT id FROM classify WHERE name = %s AND level = 1 AND parentid = %s AND id != %s
            """, (name, parent_id, classify_id))
            duplicate = cursor.fetchone()
            if duplicate:
                return jsonify({'error': '该父级类型下已存在同名的二级类型'}), 400

        # 更新类型
        cursor.execute("""
            UPDATE classify
            SET name = %s, level = %s, parentid = %s
            WHERE id = %s
        """, (name, level, parent_id, classify_id))

        connection.commit()

        return jsonify({'message': '类型更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新类型状态
@app.route('/api/classifies/<int:classify_id>/status', methods=['PUT'])
def update_classify_status(classify_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status not in [0, 1]:
            return jsonify({'error': '状态值必须为0或1'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新类型状态
        cursor.execute("""
            UPDATE classify
            SET status = %s
            WHERE id = %s
        """, (status, classify_id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '类型不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 菜单管理API ====================

@app.route('/api/menus', methods=['GET'])
def get_menus():
    """获取菜单树结构"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取所有菜单项，按sort_order排序
        cursor.execute("""
            SELECT id, name, parent_id, route, sort_order
            FROM menu
            ORDER BY sort_order ASC
        """)
        all_menus = cursor.fetchall()

        # 构建树形结构
        menu_tree = []
        menu_map = {}

        # 先创建所有菜单项的映射
        for menu in all_menus:
            menu_map[menu['id']] = {
                'id': menu['id'],
                'name': menu['name'],
                'route': menu['route'],
                'parent_id': menu['parent_id'],
                'sort_order': menu['sort_order'],
                'children': []
            }

        # 构建父子关系
        for menu in all_menus:
            if menu['parent_id'] is None:
                # 一级菜单
                menu_tree.append(menu_map[menu['id']])
            else:
                # 二级菜单，添加到父菜单的children中
                if menu['parent_id'] in menu_map:
                    menu_map[menu['parent_id']]['children'].append(menu_map[menu['id']])

        return jsonify({'menus': menu_tree}), 200

    except Exception as e:
        print(f"获取菜单失败: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 品牌管理API ====================

# API接口：获取品牌列表
@app.route('/api/brands', methods=['GET'])
def get_brands():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询所有品牌
        cursor.execute("SELECT id, name, status FROM brand ORDER BY id")
        brands = cursor.fetchall()

        return jsonify({'brands': brands}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增品牌
@app.route('/api/brands', methods=['POST'])
def add_brand():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')

        if not name:
            return jsonify({'error': '品牌名称不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查品牌名称是否已存在
        cursor.execute("SELECT id FROM brand WHERE name = %s", (name,))
        existing_brand = cursor.fetchone()
        if existing_brand:
            return jsonify({'error': '该品牌名称已存在'}), 400

        # 插入新品牌
        cursor.execute("INSERT INTO brand (name, status) VALUES (%s, 0)", (name,))
        connection.commit()

        return jsonify({'message': '品牌添加成功'}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新品牌信息
@app.route('/api/brands/<int:brand_id>', methods=['PUT'])
def update_brand(brand_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')

        if not name:
            return jsonify({'error': '品牌名称不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查品牌是否存在
        cursor.execute("SELECT id FROM brand WHERE id = %s", (brand_id,))
        brand = cursor.fetchone()
        if not brand:
            return jsonify({'error': '品牌不存在'}), 404

        # 检查品牌名称是否已被其他品牌使用
        cursor.execute("SELECT id FROM brand WHERE name = %s AND id != %s", (name, brand_id))
        existing_brand = cursor.fetchone()
        if existing_brand:
            return jsonify({'error': '该品牌名称已存在'}), 400

        # 更新品牌信息
        cursor.execute("UPDATE brand SET name = %s WHERE id = %s", (name, brand_id))
        connection.commit()

        return jsonify({'message': '品牌信息更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新品牌状态
@app.route('/api/brands/<int:brand_id>/status', methods=['PUT'])
def update_brand_status(brand_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status is None:
            return jsonify({'error': '状态不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查品牌是否存在
        cursor.execute("SELECT id FROM brand WHERE id = %s", (brand_id,))
        brand = cursor.fetchone()
        if not brand:
            return jsonify({'error': '品牌不存在'}), 404

        # 更新品牌状态
        cursor.execute("UPDATE brand SET status = %s WHERE id = %s", (status, brand_id))
        connection.commit()

        return jsonify({'message': '品牌状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 商品管理API（基础功能）====================

# API接口：获取商品列表
@app.route('/api/goods', methods=['GET'])
def get_goods():
    connection = None
    cursor = None
    try:
        # 获取查询参数
        classifyid = request.args.get('classifyid', type=int)
        status = request.args.get('status', type=int)

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 构建查询SQL
        sql = """
            SELECT
                g.id, g.name, g.brandid, g.classifyid, g.isgroup, g.price, g.status,
                b.name as brand_name,
                c.name as classify_name
            FROM goods g
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify c ON g.classifyid = c.id
            WHERE 1=1
        """
        params = []

        if classifyid is not None:
            sql += " AND g.classifyid = %s"
            params.append(classifyid)

        if status is not None:
            sql += " AND g.status = %s"
            params.append(status)

        sql += " ORDER BY g.id DESC"

        cursor.execute(sql, params)
        goods_list = cursor.fetchall()

        # 为每个商品获取属性值信息
        for goods in goods_list:
            cursor.execute("""
                SELECT
                    av.id, av.name as value_name, av.attributeid,
                    a.name as attr_name, a.classify
                FROM goods_attributevalue gav
                JOIN attribute_value av ON gav.attributevalueid = av.id
                JOIN attribute a ON av.attributeid = a.id
                WHERE gav.goodsid = %s
            """, (goods['id'],))
            attr_values = cursor.fetchall()

            # 格式化属性值显示，区分属性和规格
            if attr_values:
                # 分别收集属性（classify=0）和规格（classify=1）
                attributes_dict = {}  # 属性
                specs_dict = {}  # 规格

                for av in attr_values:
                    attr_name = av['attr_name']
                    if av['classify'] == 0:  # 属性
                        if attr_name not in attributes_dict:
                            attributes_dict[attr_name] = []
                        attributes_dict[attr_name].append(av['value_name'])
                    elif av['classify'] == 1:  # 规格
                        if attr_name not in specs_dict:
                            specs_dict[attr_name] = []
                        specs_dict[attr_name].append(av['value_name'])

                # 拼接成字符串：属性:属性值,规格:规格值
                parts = []
                for k, v in attributes_dict.items():
                    parts.append(f"{k}:{'/'.join(v)}")
                for k, v in specs_dict.items():
                    parts.append(f"{k}:{'/'.join(v)}")

                goods['attributes'] = ','.join(parts) if parts else ''

                # 保存完整的属性列表用于悬浮显示
                goods['attributes_full'] = parts
            else:
                goods['attributes'] = ''
                goods['attributes_full'] = []

        return jsonify({'goods': goods_list}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：新增商品
@app.route('/api/goods', methods=['POST'])
def add_goods():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        brandid = data.get('brandid')
        classifyid = data.get('classifyid')
        isgroup = data.get('isgroup', 1)  # 0=组合售卖，1=非组合售卖，默认为非组合
        price = data.get('price')
        attribute_values = data.get('attributevalue_ids', [])  # 属性值ID列表
        included_goods_ids = data.get('included_goods_ids', [])  # 包含商品ID列表

        if not name or not brandid or not classifyid or not price:
            return jsonify({'error': '商品信息不完整'}), 400

        # 验证：组合售卖=是时，必须至少添加一个包含商品
        if isgroup == 0 and (not included_goods_ids or len(included_goods_ids) == 0):
            return jsonify({'error': '组合商品必须至少包含一个子商品'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 插入商品
        cursor.execute("""
            INSERT INTO goods (name, brandid, classifyid, isgroup, price, status)
            VALUES (%s, %s, %s, %s, %s, 0)
        """, (name, brandid, classifyid, isgroup, price))

        goods_id = cursor.lastrowid

        # 插入商品属性值关系
        if attribute_values:
            for av_id in attribute_values:
                cursor.execute("""
                    INSERT INTO goods_attributevalue (goodsid, attributevalueid)
                    VALUES (%s, %s)
                """, (goods_id, av_id))

        # 插入商品组合关系
        if isgroup == 0 and included_goods_ids:
            for sub_goods_id in included_goods_ids:
                cursor.execute("""
                    INSERT INTO goods_goods (goodsid, parentsid)
                    VALUES (%s, %s)
                """, (sub_goods_id, goods_id))

        connection.commit()
        return jsonify({'message': '商品添加成功', 'id': goods_id}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取单个商品详情
@app.route('/api/goods/<int:goods_id>', methods=['GET'])
def get_goods_detail(goods_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询商品基本信息
        cursor.execute("""
            SELECT
                g.id, g.name, g.brandid, g.classifyid, g.isgroup, g.price, g.status,
                b.name as brand_name,
                c.name as classify_name
            FROM goods g
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify c ON g.classifyid = c.id
            WHERE g.id = %s
        """, (goods_id,))
        goods = cursor.fetchone()

        if not goods:
            return jsonify({'error': '商品不存在'}), 404

        # 查询商品的属性值
        cursor.execute("""
            SELECT attributevalueid
            FROM goods_attributevalue
            WHERE goodsid = %s
        """, (goods_id,))
        attr_values = cursor.fetchall()
        goods['attributevalue_ids'] = [av['attributevalueid'] for av in attr_values]

        # 查询包含的子商品ID
        cursor.execute("""
            SELECT goodsid
            FROM goods_goods
            WHERE parentsid = %s
            ORDER BY goodsid
        """, (goods_id,))
        included_goods = cursor.fetchall()
        goods['included_goods_ids'] = [ig['goodsid'] for ig in included_goods]

        return jsonify({'goods': goods}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新商品信息
@app.route('/api/goods/<int:goods_id>', methods=['PUT'])
def update_goods(goods_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        brandid = data.get('brandid')
        classifyid = data.get('classifyid')
        price = data.get('price')
        attribute_values = data.get('attributevalue_ids', [])
        included_goods_ids = data.get('included_goods_ids', [])  # 包含商品ID列表

        if not name or not brandid or not classifyid or not price:
            return jsonify({'error': '商品信息不完整'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查商品是否存在，并获取 isgroup
        cursor.execute("SELECT id, isgroup FROM goods WHERE id = %s", (goods_id,))
        existing_goods = cursor.fetchone()
        if not existing_goods:
            return jsonify({'error': '商品不存在'}), 404

        isgroup = existing_goods['isgroup']

        # 验证：组合售卖=是时，必须至少添加一个包含商品
        if isgroup == 0 and (not included_goods_ids or len(included_goods_ids) == 0):
            return jsonify({'error': '组合商品必须至少包含一个子商品'}), 400

        # 更新商品基本信息（不更新 isgroup）
        cursor.execute("""
            UPDATE goods
            SET name = %s, brandid = %s, classifyid = %s, price = %s
            WHERE id = %s
        """, (name, brandid, classifyid, price, goods_id))

        # 删除旧的属性值关系
        cursor.execute("DELETE FROM goods_attributevalue WHERE goodsid = %s", (goods_id,))

        # 插入新的属性值关系
        if attribute_values:
            for av_id in attribute_values:
                cursor.execute("""
                    INSERT INTO goods_attributevalue (goodsid, attributevalueid)
                    VALUES (%s, %s)
                """, (goods_id, av_id))

        # 删除旧的商品组合关系
        cursor.execute("DELETE FROM goods_goods WHERE parentsid = %s", (goods_id,))

        # 插入新的商品组合关系
        if isgroup == 0 and included_goods_ids:
            for sub_goods_id in included_goods_ids:
                cursor.execute("""
                    INSERT INTO goods_goods (goodsid, parentsid)
                    VALUES (%s, %s)
                """, (sub_goods_id, goods_id))

        connection.commit()
        return jsonify({'message': '商品信息更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新商品状态
@app.route('/api/goods/<int:goods_id>/status', methods=['PUT'])
def update_goods_status(goods_id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status is None:
            return jsonify({'error': '状态不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查商品是否存在
        cursor.execute("SELECT id FROM goods WHERE id = %s", (goods_id,))
        if not cursor.fetchone():
            return jsonify({'error': '商品不存在'}), 404

        # 更新商品状态
        cursor.execute("UPDATE goods SET status = %s WHERE id = %s", (status, goods_id))
        connection.commit()

        return jsonify({'message': '商品状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的品牌列表（用于下拉框）
@app.route('/api/brands/active', methods=['GET'])
def get_active_brands():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT id, name FROM brand WHERE status = 0 ORDER BY id")
        brands = cursor.fetchall()
        return jsonify({'brands': brands}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的类型列表（用于下拉框）
@app.route('/api/classifies/active', methods=['GET'])
def get_active_classifies():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)
        cursor.execute("SELECT id, name FROM classify WHERE status = 0 AND level = 1 ORDER BY id")
        classifies = cursor.fetchall()
        return jsonify({'classifies': classifies}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用的属性列表（用于下拉框）
@app.route('/api/attributes/active', methods=['GET'])
def get_active_attributes():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取启用的属性（包含classify字段用于区分属性和规格）
        cursor.execute("SELECT id, name, classify FROM attribute WHERE status = 0 ORDER BY id")
        attributes = cursor.fetchall()

        # 为每个属性获取其属性值
        for attribute in attributes:
            cursor.execute("""
                SELECT id, name
                FROM attribute_value
                WHERE attributeid = %s
                ORDER BY id
            """, (attribute['id'],))
            attribute['values'] = cursor.fetchall()

        return jsonify({'attributes': attributes}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取可用于组合的商品列表（用于下拉框）
@app.route('/api/goods/available-for-combo', methods=['GET'])
def get_available_for_combo():
    connection = None
    cursor = None
    try:
        exclude_id = request.args.get('exclude_id', type=int)

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询启用且非组合的商品
        sql = """
            SELECT
                g.id, g.name, g.brandid, g.classifyid, g.price,
                b.name as brand_name,
                c.name as classify_name
            FROM goods g
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify c ON g.classifyid = c.id
            WHERE g.status = 0 AND g.isgroup = 1
        """

        params = []
        if exclude_id:
            sql += " AND g.id != %s"
            params.append(exclude_id)

        sql += " ORDER BY g.id DESC"

        cursor.execute(sql, params)
        goods_list = cursor.fetchall()

        # 为每个商品获取属性值信息（格式化为 "属性:值,属性:值"）
        for goods in goods_list:
            cursor.execute("""
                SELECT
                    a.name as attr_name,
                    av.name as value_name
                FROM goods_attributevalue gav
                JOIN attribute_value av ON gav.attributevalueid = av.id
                JOIN attribute a ON av.attributeid = a.id
                WHERE gav.goodsid = %s
                ORDER BY a.classify, a.id
            """, (goods['id'],))
            attr_values = cursor.fetchall()

            # 格式化为 "颜色:黑色,内存:256GB"
            if attr_values:
                attrs_dict = {}
                for av in attr_values:
                    attr_name = av['attr_name']
                    if attr_name not in attrs_dict:
                        attrs_dict[attr_name] = []
                    attrs_dict[attr_name].append(av['value_name'])

                attrs_parts = [f"{k}:{'/'.join(v)}" for k, v in attrs_dict.items()]
                goods['attributes'] = ','.join(attrs_parts)
            else:
                goods['attributes'] = ''

        return jsonify({'goods': goods_list}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取组合商品包含的子商品列表
@app.route('/api/goods/<int:goods_id>/included-goods', methods=['GET'])
def get_included_goods(goods_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询包含的商品列表
        cursor.execute("""
            SELECT
                g.id, g.name, g.price,
                b.name as brand_name,
                c.name as classify_name
            FROM goods_goods gg
            JOIN goods g ON gg.goodsid = g.id
            LEFT JOIN brand b ON g.brandid = b.id
            LEFT JOIN classify c ON g.classifyid = c.id
            WHERE gg.parentsid = %s
            ORDER BY gg.goodsid
        """, (goods_id,))

        included_goods = cursor.fetchall()

        # 为每个商品获取属性值信息
        for goods in included_goods:
            cursor.execute("""
                SELECT
                    a.name as attr_name,
                    av.name as value_name
                FROM goods_attributevalue gav
                JOIN attribute_value av ON gav.attributevalueid = av.id
                JOIN attribute a ON av.attributeid = a.id
                WHERE gav.goodsid = %s
                ORDER BY a.classify, a.id
            """, (goods['id'],))
            attr_values = cursor.fetchall()

            if attr_values:
                attrs_dict = {}
                for av in attr_values:
                    attr_name = av['attr_name']
                    if attr_name not in attrs_dict:
                        attrs_dict[attr_name] = []
                    attrs_dict[attr_name].append(av['value_name'])

                attrs_parts = [f"{k}:{'/'.join(v)}" for k, v in attrs_dict.items()]
                goods['attributes'] = ','.join(attrs_parts)
            else:
                goods['attributes'] = ''

        return jsonify({'included_goods': included_goods}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 活动模板API ====================

# API接口：获取活动模板列表
@app.route('/api/activity-templates', methods=['GET'])
def get_activity_templates():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        cursor.execute("SELECT * FROM activity_template ORDER BY id DESC")
        templates = cursor.fetchall()

        return jsonify({'templates': templates}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取单个活动模板详情
@app.route('/api/activity-templates/<int:id>', methods=['GET'])
def get_activity_template(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取模板基本信息
        cursor.execute("SELECT * FROM activity_template WHERE id = %s", (id,))
        template = cursor.fetchone()

        if not template:
            return jsonify({'error': '活动模板不存在'}), 404

        # 获取关联的商品或类型
        if template['select_type'] == 1:  # 按类型
            cursor.execute("""
                SELECT atg.classify_id, c.name as classify_name
                FROM activity_template_goods atg
                JOIN classify c ON atg.classify_id = c.id
                WHERE atg.template_id = %s
            """, (id,))
            template['classify_list'] = cursor.fetchall()
        else:  # 按商品
            cursor.execute("""
                SELECT atg.goods_id, g.name as goods_name, g.price, b.name as brand_name, c.name as classify_name
                FROM activity_template_goods atg
                JOIN goods g ON atg.goods_id = g.id
                LEFT JOIN brand b ON g.brandid = b.id
                LEFT JOIN classify c ON g.classifyid = c.id
                WHERE atg.template_id = %s
            """, (id,))
            template['goods_list'] = cursor.fetchall()

        return jsonify({'template': template}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：创建活动模板
@app.route('/api/activity-templates', methods=['POST'])
def create_activity_template():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        type_val = data.get('type')
        select_type = data.get('select_type')
        status = data.get('status', 0)
        classify_ids = data.get('classify_ids', [])
        goods_ids = data.get('goods_ids', [])

        if not name or not type_val or not select_type:
            return jsonify({'error': '模板名称、活动类型和选择方式不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 创建活动模板
        cursor.execute("""
            INSERT INTO activity_template (name, type, select_type, status)
            VALUES (%s, %s, %s, %s)
        """, (name, type_val, select_type, status))

        template_id = cursor.lastrowid

        # 插入关联数据
        if select_type == 1 and classify_ids:  # 按类型
            for classify_id in classify_ids:
                cursor.execute("""
                    INSERT INTO activity_template_goods (template_id, classify_id)
                    VALUES (%s, %s)
                """, (template_id, classify_id))
        elif select_type == 2 and goods_ids:  # 按商品
            for goods_id in goods_ids:
                cursor.execute("""
                    INSERT INTO activity_template_goods (template_id, goods_id)
                    VALUES (%s, %s)
                """, (template_id, goods_id))

        connection.commit()

        return jsonify({'message': '活动模板创建成功', 'id': template_id}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新活动模板
@app.route('/api/activity-templates/<int:id>', methods=['PUT'])
def update_activity_template(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        type_val = data.get('type')
        select_type = data.get('select_type')
        status = data.get('status', 0)
        classify_ids = data.get('classify_ids', [])
        goods_ids = data.get('goods_ids', [])

        if not name or not type_val or not select_type:
            return jsonify({'error': '模板名称、活动类型和选择方式不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新活动模板
        cursor.execute("""
            UPDATE activity_template
            SET name = %s, type = %s, select_type = %s, status = %s
            WHERE id = %s
        """, (name, type_val, select_type, status, id))

        # 删除原有关联数据
        cursor.execute("DELETE FROM activity_template_goods WHERE template_id = %s", (id,))

        # 插入新的关联数据
        if select_type == 1 and classify_ids:  # 按类型
            for classify_id in classify_ids:
                cursor.execute("""
                    INSERT INTO activity_template_goods (template_id, classify_id)
                    VALUES (%s, %s)
                """, (id, classify_id))
        elif select_type == 2 and goods_ids:  # 按商品
            for goods_id in goods_ids:
                cursor.execute("""
                    INSERT INTO activity_template_goods (template_id, goods_id)
                    VALUES (%s, %s)
                """, (id, goods_id))

        connection.commit()

        return jsonify({'message': '活动模板更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：删除活动模板
@app.route('/api/activity-templates/<int:id>', methods=['DELETE'])
def delete_activity_template(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查是否有关联活动
        cursor.execute("SELECT COUNT(*) as count FROM activity WHERE template_id = %s", (id,))
        result = cursor.fetchone()
        if result['count'] > 0:
            return jsonify({'error': '该模板有关联活动，无法删除'}), 400

        # 删除活动模板（关联数据会级联删除）
        cursor.execute("DELETE FROM activity_template WHERE id = %s", (id,))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '活动模板不存在'}), 404

        return jsonify({'message': '活动模板删除成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新活动模板状态
@app.route('/api/activity-templates/<int:id>/status', methods=['PUT'])
def update_activity_template_status(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status is None:
            return jsonify({'error': '状态不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("UPDATE activity_template SET status = %s WHERE id = %s", (status, id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '活动模板不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 活动管理API ====================

# API接口：获取活动列表
@app.route('/api/activities', methods=['GET'])
def get_activities():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        cursor.execute("""
            SELECT a.*, t.name as template_name, t.type as template_type
            FROM activity a
            JOIN activity_template t ON a.template_id = t.id
            ORDER BY a.id DESC
        """)
        activities = cursor.fetchall()

        return jsonify({'activities': activities}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取单个活动详情
@app.route('/api/activities/<int:id>', methods=['GET'])
def get_activity(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取活动基本信息
        cursor.execute("""
            SELECT a.*, t.name as template_name, t.type as template_type, t.select_type
            FROM activity a
            JOIN activity_template t ON a.template_id = t.id
            WHERE a.id = %s
        """, (id,))
        activity = cursor.fetchone()

        if not activity:
            return jsonify({'error': '活动不存在'}), 404

        # 获取活动明细（满折规则）
        cursor.execute("""
            SELECT * FROM activity_detail WHERE activity_id = %s ORDER BY threshold_amount
        """, (id,))
        activity['details'] = cursor.fetchall()

        return jsonify({'activity': activity}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：创建活动
@app.route('/api/activities', methods=['POST'])
def create_activity():
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        template_id = data.get('template_id')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        status = data.get('status', 0)
        details = data.get('details', [])

        if not name or not template_id or not start_time or not end_time:
            return jsonify({'error': '活动名称、关联模板、开始时间和结束时间不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 创建活动
        cursor.execute("""
            INSERT INTO activity (name, template_id, start_time, end_time, status)
            VALUES (%s, %s, %s, %s, %s)
        """, (name, template_id, start_time, end_time, status))

        activity_id = cursor.lastrowid

        # 插入活动明细
        for detail in details:
            cursor.execute("""
                INSERT INTO activity_detail (activity_id, threshold_amount, discount_value)
                VALUES (%s, %s, %s)
            """, (activity_id, detail['threshold_amount'], detail['discount_value']))

        connection.commit()

        return jsonify({'message': '活动创建成功', 'id': activity_id}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新活动
@app.route('/api/activities/<int:id>', methods=['PUT'])
def update_activity(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        name = data.get('name')
        template_id = data.get('template_id')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        status = data.get('status', 0)
        details = data.get('details', [])

        if not name or not template_id or not start_time or not end_time:
            return jsonify({'error': '活动名称、关联模板、开始时间和结束时间不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        # 更新活动
        cursor.execute("""
            UPDATE activity
            SET name = %s, template_id = %s, start_time = %s, end_time = %s, status = %s
            WHERE id = %s
        """, (name, template_id, start_time, end_time, status, id))

        # 删除原有明细
        cursor.execute("DELETE FROM activity_detail WHERE activity_id = %s", (id,))

        # 插入新的明细
        for detail in details:
            cursor.execute("""
                INSERT INTO activity_detail (activity_id, threshold_amount, discount_value)
                VALUES (%s, %s, %s)
            """, (id, detail['threshold_amount'], detail['discount_value']))

        connection.commit()

        return jsonify({'message': '活动更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：删除活动
@app.route('/api/activities/<int:id>', methods=['DELETE'])
def delete_activity(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        # 删除活动（明细会级联删除）
        cursor.execute("DELETE FROM activity WHERE id = %s", (id,))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '活动不存在'}), 404

        return jsonify({'message': '活动删除成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：更新活动状态
@app.route('/api/activities/<int:id>/status', methods=['PUT'])
def update_activity_status(id):
    connection = None
    cursor = None
    try:
        data = request.get_json()
        status = data.get('status')

        if status is None:
            return jsonify({'error': '状态不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor()

        cursor.execute("UPDATE activity SET status = %s WHERE id = %s", (status, id))

        connection.commit()

        if cursor.rowcount == 0:
            return jsonify({'error': '活动不存在'}), 404

        return jsonify({'message': '状态更新成功'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：获取启用状态的活动模板列表
@app.route('/api/activity-templates/active', methods=['GET'])
def get_active_activity_templates():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        cursor.execute("SELECT * FROM activity_template WHERE status = 0 ORDER BY id DESC")
        templates = cursor.fetchall()

        return jsonify({'templates': templates}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：根据预计付款时间查询启用的活动
@app.route('/api/activities/by-date-range', methods=['GET'])
def get_activities_by_date_range():
    """根据预计付款时间查询该时间范围内status=0（启用）的活动"""
    connection = None
    cursor = None
    try:
        payment_time = request.args.get('payment_time')

        if not payment_time:
            return jsonify({'error': '预计付款时间不能为空'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询在预计付款时间范围内且启用的活动
        cursor.execute("""
            SELECT
                a.id,
                a.name,
                a.template_id,
                a.start_time,
                a.end_time,
                a.status,
                t.name AS template_name,
                t.type AS template_type,
                t.select_type AS template_select_type
            FROM activity a
            JOIN activity_template t ON a.template_id = t.id
            WHERE a.start_time <= %s
              AND a.end_time >= %s
              AND a.status = 0
            ORDER BY a.id ASC
        """, (payment_time, payment_time))

        activities = cursor.fetchall()

        # 检测同类型活动重复
        type_count = {}
        duplicate_type = None
        has_duplicate = False

        for activity in activities:
            activity_type = activity['template_type']
            if activity_type in type_count:
                has_duplicate = True
                duplicate_type = activity_type
                break
            type_count[activity_type] = 1

        # 如果有重复，返回错误信息
        if has_duplicate:
            type_names = {1: '满减', 2: '满折', 3: '满赠', 4: '换购'}
            return jsonify({
                'has_duplicate': True,
                'duplicate_type': duplicate_type,
                'type_name': type_names.get(duplicate_type, '未知'),
                'activities': []
            }), 200

        # 为每个活动查询满折规则（仅type=2的活动）
        for activity in activities:
            if activity['template_type'] == 2:  # 满折类型
                cursor.execute("""
                    SELECT threshold_amount, discount_value
                    FROM activity_detail
                    WHERE activity_id = %s
                    ORDER BY threshold_amount ASC
                """, (activity['id'],))
                activity['details'] = cursor.fetchall()
            else:
                activity['details'] = []

        return jsonify({
            'has_duplicate': False,
            'duplicate_type': None,
            'activities': activities
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# API接口：计算订单优惠金额
@app.route('/api/orders/calculate-discount', methods=['POST'])
def calculate_order_discount():
    """计算订单优惠金额（满折）"""
    connection = None
    cursor = None
    try:
        data = request.get_json()
        goods_list = data.get('goods_list', [])  # [{goods_id, price, total_price}]
        activity_ids = data.get('activity_ids', [])

        if not goods_list:
            return jsonify({'total_discount': 0, 'child_discounts': {}}), 200

        if not activity_ids:
            return jsonify({'total_discount': 0, 'child_discounts': {}}), 200

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        total_discount = 0
        all_child_discounts = {}

        # 初始化每个商品的优惠为0
        for goods in goods_list:
            all_child_discounts[str(goods['goods_id'])] = 0

        # 遍历每个活动计算优惠
        for activity_id in activity_ids:
            # 查询活动信息
            cursor.execute("""
                SELECT a.id, a.template_id, t.type, t.select_type
                FROM activity a
                JOIN activity_template t ON a.template_id = t.id
                WHERE a.id = %s
            """, (activity_id,))

            activity = cursor.fetchone()
            if not activity:
                continue

            # 仅处理满折类型（type=2）
            if activity['type'] != 2:
                continue

            # 查询活动满折规则
            cursor.execute("""
                SELECT threshold_amount, discount_value
                FROM activity_detail
                WHERE activity_id = %s
                ORDER BY threshold_amount DESC
            """, (activity_id,))

            details = cursor.fetchall()
            if not details:
                continue

            # 查询活动模板关联的商品/类型
            cursor.execute("""
                SELECT goods_id, classify_id
                FROM activity_template_goods
                WHERE template_id = %s
            """, (activity['template_id'],))

            template_goods = cursor.fetchall()
            if not template_goods:
                continue

            # 筛选参与活动的商品
            eligible_goods = []
            for goods in goods_list:
                is_eligible = False

                if activity['select_type'] == 2:  # 按商品
                    # 检查商品ID是否在模板关联的商品中
                    for tg in template_goods:
                        if tg['goods_id'] == goods['goods_id']:
                            is_eligible = True
                            break
                else:  # 按类型
                    # 查询商品的classifyid
                    cursor.execute("SELECT classifyid FROM goods WHERE id = %s", (goods['goods_id'],))
                    goods_info = cursor.fetchone()
                    if goods_info:
                        for tg in template_goods:
                            if tg['classify_id'] == goods_info['classifyid']:
                                is_eligible = True
                                break

                if is_eligible:
                    eligible_goods.append(goods)

            if not eligible_goods:
                continue

            # 计算参与商品数量
            eligible_count = len(eligible_goods)

            # 匹配最大满足的折扣档位
            matched_discount = None
            for detail in details:
                if eligible_count >= float(detail['threshold_amount']):
                    matched_discount = float(detail['discount_value'])
                    break

            if not matched_discount:
                continue

            # 计算总优惠：(1 - 折扣/100) × 参与商品标准售价之和
            # discount_value=90表示9折(付90%)，discount_value=80表示8折(付80%)
            eligible_price_sum = sum(float(g['price']) for g in eligible_goods)
            discount_rate = 1 - matched_discount / 100  # 80 -> 1 - 0.8 = 0.2 (优惠20%)
            activity_discount = round(discount_rate * eligible_price_sum, 2)

            total_discount += activity_discount

            # 按比例分摊优惠到各商品
            if activity_discount > 0 and eligible_price_sum > 0:
                allocated = 0
                for i, goods in enumerate(eligible_goods):
                    if i == len(eligible_goods) - 1:
                        # 最后一个商品用减法避免精度误差
                        child_discount = activity_discount - allocated
                    else:
                        ratio = float(goods['price']) / eligible_price_sum
                        child_discount = round(activity_discount * ratio, 2)
                        allocated += child_discount

                    goods_id_str = str(goods['goods_id'])
                    all_child_discounts[goods_id_str] += child_discount

        # 四舍五入所有子订单优惠
        for goods_id in all_child_discounts:
            all_child_discounts[goods_id] = round(all_child_discounts[goods_id], 2)

        return jsonify({
            'total_discount': round(total_discount, 2),
            'child_discounts': all_child_discounts
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 合同管理接口 ====================

# 获取合同列表
@app.route('/api/contracts', methods=['GET'])
def get_contracts():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 构建查询SQL
        sql = """
            SELECT
                c.id, c.name, c.student_id, c.type, c.signature_form,
                c.contract_amount, c.signatory, c.initiating_party, c.initiator,
                c.status, c.payment_status, c.create_time,
                s.name AS student_name
            FROM contract c
            LEFT JOIN student s ON c.student_id = s.id
            ORDER BY c.id DESC
        """

        cursor.execute(sql)
        contracts = cursor.fetchall()

        return jsonify({'contracts': contracts}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取合同详情
@app.route('/api/contracts/<int:contract_id>', methods=['GET'])
def get_contract(contract_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        cursor.execute("""
            SELECT
                c.id, c.name, c.student_id, c.type, c.signature_form,
                c.contract_amount, c.signatory, c.initiating_party, c.initiator,
                c.status, c.payment_status, c.termination_agreement, c.create_time,
                s.name AS student_name
            FROM contract c
            LEFT JOIN student s ON c.student_id = s.id
            WHERE c.id = %s
        """, (contract_id,))

        contract = cursor.fetchone()

        if not contract:
            return jsonify({'error': '合同不存在'}), 404

        return jsonify({'contract': contract}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 新增合同
@app.route('/api/contracts', methods=['POST'])
def add_contract():
    connection = None
    cursor = None
    try:
        data = request.json
        name = data.get('name')
        student_id = data.get('student_id')
        contract_type = data.get('type')
        signature_form = data.get('signature_form')
        contract_amount = data.get('contract_amount')
        signatory = data.get('signatory')

        # 校验必填项
        if not name or not student_id or contract_type is None or signature_form is None or not contract_amount:
            return jsonify({'error': '请填写所有必填项'}), 400

        # 获取发起人（当前登录用户）
        initiator = session.get('username', '')

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 插入合同
        cursor.execute("""
            INSERT INTO contract (name, student_id, type, signature_form, contract_amount, signatory, initiator, status, payment_status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, 0, 0)
        """, (name, student_id, contract_type, signature_form, contract_amount, signatory, initiator))

        connection.commit()
        return jsonify({'message': '合同新增成功'}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 撤销合同
@app.route('/api/contracts/<int:contract_id>/revoke', methods=['PUT'])
def revoke_contract(contract_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查合同是否存在
        cursor.execute("SELECT id, status FROM contract WHERE id = %s", (contract_id,))
        contract = cursor.fetchone()

        if not contract:
            return jsonify({'error': '合同不存在'}), 404

        # 只有待审核状态的合同可以撤销
        if contract['status'] != 0:
            return jsonify({'error': '只有待审核状态的合同可以撤销'}), 400

        # 更新状态为已作废
        cursor.execute("UPDATE contract SET status = 98 WHERE id = %s", (contract_id,))
        connection.commit()

        return jsonify({'message': '合同已撤销'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 中止合作（上传中止协议）
@app.route('/api/contracts/<int:contract_id>/terminate', methods=['PUT'])
def terminate_contract(contract_id):
    connection = None
    cursor = None
    try:
        data = request.json
        termination_agreement = data.get('termination_agreement')

        if not termination_agreement:
            return jsonify({'error': '请上传中止协议'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查合同是否存在
        cursor.execute("SELECT id, status FROM contract WHERE id = %s", (contract_id,))
        contract = cursor.fetchone()

        if not contract:
            return jsonify({'error': '合同不存在'}), 404

        # 只有已通过状态的合同可以中止
        if contract['status'] != 50:
            return jsonify({'error': '只有已通过状态的合同可以中止'}), 400

        # 更新状态为协议中止，并保存中止协议路径
        cursor.execute("""
            UPDATE contract
            SET status = 99, termination_agreement = %s
            WHERE id = %s
        """, (termination_agreement, contract_id))
        connection.commit()

        return jsonify({'message': '合作已中止'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 收款管理接口 ====================

# 根据收款情况更新订单状态的辅助函数
def update_order_payment_status(cursor, order_id):
    """
    根据订单关联的收款记录更新订单状态
    - 收款金额之和 = 0: 状态为20（未支付）
    - 0 < 收款金额之和 < 实收金额: 状态为30（部分支付）
    - 收款金额之和 >= 实收金额: 状态为40（已支付）
    """
    # 获取订单的实收金额
    cursor.execute("""
        SELECT amount_received FROM orders WHERE id = %s
    """, (order_id,))
    order = cursor.fetchone()

    if not order:
        return

    amount_received = float(order['amount_received'])

    # 计算该订单状态为未核验(10)和已支付(20)的收款金额之和
    cursor.execute("""
        SELECT COALESCE(SUM(payment_amount), 0) AS total_paid
        FROM payment_collection
        WHERE order_id = %s AND status IN (10, 20)
    """, (order_id,))

    result = cursor.fetchone()
    total_paid = float(result['total_paid']) if result else 0

    # 根据收款金额确定订单状态
    if total_paid == 0:
        new_status = 20  # 未支付
    elif total_paid >= amount_received:
        new_status = 40  # 已支付
    else:
        new_status = 30  # 部分支付

    # 更新订单状态
    cursor.execute("""
        UPDATE orders SET status = %s WHERE id = %s
    """, (new_status, order_id))

# 获取收款列表
@app.route('/api/payment-collections', methods=['GET'])
def get_payment_collections():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 构建查询SQL
        sql = """
            SELECT
                pc.id, pc.order_id, pc.student_id, pc.payment_scenario, pc.payment_method,
                pc.payment_amount, pc.payer, pc.payee_entity, pc.trading_hours,
                pc.arrival_time, pc.merchant_order, pc.status, pc.create_time,
                s.name AS student_name
            FROM payment_collection pc
            LEFT JOIN student s ON pc.student_id = s.id
            WHERE 1=1
        """
        params = []

        # 获取筛选参数
        id_filter = request.args.get('id')
        student_id = request.args.get('student_id')
        order_id = request.args.get('order_id')
        payer = request.args.get('payer')
        payment_method = request.args.get('payment_method')
        trading_date = request.args.get('trading_date')
        status = request.args.get('status')

        if id_filter:
            sql += " AND pc.id = %s"
            params.append(id_filter)

        if student_id:
            sql += " AND pc.student_id = %s"
            params.append(student_id)

        if order_id:
            sql += " AND pc.order_id = %s"
            params.append(order_id)

        if payer:
            sql += " AND pc.payer = %s"
            params.append(payer)

        if payment_method is not None and payment_method != '':
            sql += " AND pc.payment_method = %s"
            params.append(payment_method)

        if trading_date:
            sql += " AND DATE(pc.trading_hours) = %s"
            params.append(trading_date)

        if status is not None and status != '':
            sql += " AND pc.status = %s"
            params.append(status)

        sql += " ORDER BY pc.id DESC"

        cursor.execute(sql, params)
        collections = cursor.fetchall()

        return jsonify({'collections': collections}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 获取学生的待支付/部分支付订单
@app.route('/api/students/<int:student_id>/unpaid-orders', methods=['GET'])
def get_student_unpaid_orders(student_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询该学生状态为未支付(20)或部分支付(30)的订单，并且实收金额大于0（表示有待支付金额）
        cursor.execute("""
            SELECT id, amount_received, expected_payment_time
            FROM orders
            WHERE student_id = %s AND status IN (20, 30) AND amount_received > 0
            ORDER BY id DESC
        """, (student_id,))

        orders = cursor.fetchall()
        return jsonify({'orders': orders}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 计算订单的待支付金额
@app.route('/api/orders/<int:order_id>/pending-amount', methods=['GET'])
def get_order_pending_amount(order_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取订单的实收金额
        cursor.execute("""
            SELECT amount_received, expected_payment_time
            FROM orders
            WHERE id = %s
        """, (order_id,))

        order = cursor.fetchone()
        if not order:
            return jsonify({'error': '订单不存在'}), 404

        # 计算该订单已有的未核验、已支付的收款金额之和
        cursor.execute("""
            SELECT COALESCE(SUM(payment_amount), 0) AS paid_amount
            FROM payment_collection
            WHERE order_id = %s AND status IN (10, 20)
        """, (order_id,))

        result = cursor.fetchone()
        paid_amount = float(result['paid_amount']) if result else 0

        # 待支付金额 = 实收金额 - 已付金额
        pending_amount = float(order['amount_received']) - paid_amount

        return jsonify({
            'pending_amount': round(pending_amount, 2),
            'amount_received': float(order['amount_received']),
            'paid_amount': round(paid_amount, 2),
            'expected_payment_time': order['expected_payment_time']
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 新增收款
@app.route('/api/payment-collections', methods=['POST'])
def add_payment_collection():
    connection = None
    cursor = None
    try:
        data = request.json
        order_id = data.get('order_id')
        student_id = data.get('student_id')
        payment_scenario = data.get('payment_scenario')
        payment_method = data.get('payment_method')
        payment_amount = data.get('payment_amount')
        payer = data.get('payer')
        payee_entity = data.get('payee_entity')
        trading_hours = data.get('trading_hours')
        merchant_order = data.get('merchant_order')

        # 校验必填项
        if not order_id or not student_id or payment_scenario is None or payment_method is None or not payment_amount:
            return jsonify({'error': '请填写所有必填项'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取订单的实收金额
        cursor.execute("""
            SELECT amount_received
            FROM orders
            WHERE id = %s
        """, (order_id,))

        order = cursor.fetchone()
        if not order:
            return jsonify({'error': '订单不存在'}), 404

        # 计算该订单已有的未核验、已支付的收款金额之和
        cursor.execute("""
            SELECT COALESCE(SUM(payment_amount), 0) AS paid_amount
            FROM payment_collection
            WHERE order_id = %s AND status IN (10, 20)
        """, (order_id,))

        result = cursor.fetchone()
        paid_amount = float(result['paid_amount']) if result else 0

        # 待支付金额
        pending_amount = float(order['amount_received']) - paid_amount

        # 校验：付款金额不能超过待支付金额
        if float(payment_amount) > pending_amount:
            return jsonify({'error': f'付款金额不能超过待支付金额({pending_amount})'}), 400

        # 根据付款场景设置初始状态
        # 线上(0) -> 待支付(0)
        # 线下(1) -> 未核验(10)
        initial_status = 0 if payment_scenario == 0 else 10

        # 插入收款记录
        cursor.execute("""
            INSERT INTO payment_collection (order_id, student_id, payment_scenario, payment_method,
                payment_amount, payer, payee_entity, trading_hours, merchant_order, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (order_id, student_id, payment_scenario, payment_method,
              payment_amount, payer, payee_entity, trading_hours, merchant_order, initial_status))

        # 更新订单状态
        update_order_payment_status(cursor, order_id)

        connection.commit()
        return jsonify({'message': '收款新增成功'}), 201
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 确认到账
@app.route('/api/payment-collections/<int:collection_id>/confirm', methods=['PUT'])
def confirm_payment_collection(collection_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查收款记录是否存在
        cursor.execute("SELECT id, status, order_id, payment_amount FROM payment_collection WHERE id = %s", (collection_id,))
        collection = cursor.fetchone()

        if not collection:
            return jsonify({'error': '收款记录不存在'}), 404

        # 只有未核验状态可以确认到账
        if collection['status'] != 10:
            return jsonify({'error': '只有未核验状态的收款可以确认到账'}), 400

        # 更新状态为已支付，设置到账时间
        cursor.execute("""
            UPDATE payment_collection
            SET status = 20, arrival_time = NOW()
            WHERE id = %s
        """, (collection_id,))

        # 更新订单状态
        order_id = collection['order_id']
        update_order_payment_status(cursor, order_id)

        connection.commit()
        return jsonify({'message': '已确认到账'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 删除收款
@app.route('/api/payment-collections/<int:collection_id>', methods=['DELETE'])
def delete_payment_collection(collection_id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查收款记录是否存在
        cursor.execute("SELECT id, status, order_id FROM payment_collection WHERE id = %s", (collection_id,))
        collection = cursor.fetchone()

        if not collection:
            return jsonify({'error': '收款记录不存在'}), 404

        # 只有未核验状态可以删除
        if collection['status'] != 10:
            return jsonify({'error': '只有未核验状态的收款可以删除'}), 400

        # 获取订单ID
        order_id = collection['order_id']

        # 删除记录
        cursor.execute("DELETE FROM payment_collection WHERE id = %s", (collection_id,))

        # 更新订单状态
        update_order_payment_status(cursor, order_id)

        connection.commit()

        return jsonify({'message': '收款已删除'}), 200
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 临时API：修复活动折扣值
@app.route('/api/fix-activity-discount/<int:activity_id>', methods=['PUT'])
def fix_activity_discount(activity_id):
    """修复活动折扣值：将个位数转换为正确的百分比格式"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 查询需要修复的记录（discount_value < 10）
        cursor.execute("""
            SELECT id, discount_value
            FROM activity_detail
            WHERE activity_id = %s AND discount_value < 10
        """, (activity_id,))

        records = cursor.fetchall()
        if not records:
            return jsonify({'message': '没有需要修复的数据'}), 200

        # 修复每条记录：乘以10
        for record in records:
            new_value = float(record['discount_value']) * 10
            cursor.execute("""
                UPDATE activity_detail
                SET discount_value = %s
                WHERE id = %s
            """, (new_value, record['id']))

        connection.commit()

        return jsonify({
            'message': '修复成功',
            'fixed_count': len(records),
            'details': [{'id': r['id'], 'old': r['discount_value'], 'new': float(r['discount_value']) * 10} for r in records]
        }), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# ==================== 待认领收款管理接口 ====================

# 获取待认领列表
@app.route('/api/unclaimed', methods=['GET'])
def get_unclaimed_list():
    """获取待认领收款列表"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取筛选参数
        unclaimed_id = request.args.get('id', '')
        payer = request.args.get('payer', '')
        payment_method = request.args.get('payment_method', '')
        arrival_date = request.args.get('arrival_date', '')
        claimer = request.args.get('claimer', '')
        status = request.args.get('status', '')

        # 构建SQL查询
        sql = """
            SELECT
                u.id,
                u.payment_method,
                u.payment_amount,
                u.payer,
                u.payee_entity,
                u.arrival_time,
                u.claimer,
                u.status,
                ua.username AS claimer_name
            FROM unclaimed u
            LEFT JOIN useraccount ua ON u.claimer = ua.id
            WHERE 1=1
        """
        params = []

        # 添加筛选条件
        if unclaimed_id:
            sql += " AND u.id = %s"
            params.append(unclaimed_id)

        if payer:
            sql += " AND u.payer = %s"
            params.append(payer)

        if payment_method:
            sql += " AND u.payment_method = %s"
            params.append(payment_method)

        if arrival_date:
            sql += " AND DATE(u.arrival_time) = %s"
            params.append(arrival_date)

        if claimer:
            sql += " AND ua.username = %s"
            params.append(claimer)

        if status:
            sql += " AND u.status = %s"
            params.append(status)

        sql += " ORDER BY u.id DESC"

        cursor.execute(sql, params)
        unclaimed_list = cursor.fetchall()

        return jsonify({'unclaimed': unclaimed_list}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 认领收款
@app.route('/api/unclaimed/<int:unclaimed_id>/claim', methods=['PUT'])
def claim_unclaimed(unclaimed_id):
    """认领收款"""
    connection = None
    cursor = None
    try:
        # 检查是否登录
        if 'username' not in session:
            return jsonify({'error': '未登录'}), 401

        # 获取请求参数
        data = request.get_json()
        order_id = data.get('order_id')

        if not order_id:
            return jsonify({'error': '请输入订单ID'}), 400

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 获取当前用户ID
        cursor.execute("SELECT id FROM useraccount WHERE username = %s", (session['username'],))
        user = cursor.fetchone()
        if not user:
            return jsonify({'error': '用户不存在'}), 404

        user_id = user['id']

        # 检查待认领记录是否存在且状态为待认领
        cursor.execute("""
            SELECT payment_method, payment_amount, payer, payee_entity, merchant_order, arrival_time, status
            FROM unclaimed
            WHERE id = %s
        """, (unclaimed_id,))
        unclaimed = cursor.fetchone()

        if not unclaimed:
            return jsonify({'error': '待认领记录不存在'}), 404

        if unclaimed['status'] != 0:
            return jsonify({'error': '该记录已被认领'}), 400

        # 验证订单是否存在且状态正确
        cursor.execute("""
            SELECT student_id, amount_received, status
            FROM orders
            WHERE id = %s
        """, (order_id,))
        order = cursor.fetchone()

        if not order:
            return jsonify({'error': '订单不存在'}), 404

        # 验证订单状态是否为未支付(20)或部分支付(30)
        if order['status'] not in [20, 30]:
            return jsonify({'error': '订单状态必须为未支付或部分支付'}), 400

        # 计算该订单已有的收款金额
        cursor.execute("""
            SELECT COALESCE(SUM(payment_amount), 0) AS total_paid
            FROM payment_collection
            WHERE order_id = %s AND status IN (10, 20)
        """, (order_id,))
        result = cursor.fetchone()
        total_paid = float(result['total_paid']) if result else 0

        # 验证：待认领金额 + 已有收款金额 <= 订单实收金额
        unclaimed_amount = float(unclaimed['payment_amount'])
        order_amount_received = float(order['amount_received'])

        if total_paid + unclaimed_amount > order_amount_received:
            return jsonify({'error': f'收款金额超出订单实收金额（已收{total_paid}元，待认领{unclaimed_amount}元，订单实收{order_amount_received}元）'}), 400

        # 插入到payment_collection表
        cursor.execute("""
            INSERT INTO payment_collection (
                order_id, student_id, payment_scenario, payment_method,
                payment_amount, payer, payee_entity, merchant_order,
                trading_hours, arrival_time, status
            ) VALUES (%s, %s, 1, %s, %s, %s, %s, %s, %s, %s, 20)
        """, (
            order_id,
            order['student_id'],
            unclaimed['payment_method'],
            unclaimed['payment_amount'],
            unclaimed['payer'],
            unclaimed['payee_entity'],
            unclaimed['merchant_order'],
            unclaimed['arrival_time'],
            unclaimed['arrival_time']
        ))

        # 获取新插入的payment_collection的ID
        payment_id = cursor.lastrowid

        # 更新unclaimed为已认领状态，并关联payment_id
        cursor.execute("""
            UPDATE unclaimed
            SET status = 1, claimer = %s, payment_id = %s
            WHERE id = %s
        """, (user_id, payment_id, unclaimed_id))

        # 更新订单状态
        update_order_payment_status(cursor, order_id)

        connection.commit()

        return jsonify({'message': '认领成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 删除待认领记录
@app.route('/api/unclaimed/<int:unclaimed_id>', methods=['DELETE'])
def delete_unclaimed(unclaimed_id):
    """删除待认领记录"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 检查待认领记录是否存在且状态为待认领
        cursor.execute("SELECT status FROM unclaimed WHERE id = %s", (unclaimed_id,))
        unclaimed = cursor.fetchone()

        if not unclaimed:
            return jsonify({'error': '待认领记录不存在'}), 404

        if unclaimed['status'] != 0:
            return jsonify({'error': '仅能删除待认领状态的记录'}), 400

        # 删除记录
        cursor.execute("DELETE FROM unclaimed WHERE id = %s", (unclaimed_id,))
        connection.commit()

        return jsonify({'message': '删除成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 下载待认领Excel模板
@app.route('/api/unclaimed/template', methods=['GET'])
def download_unclaimed_template():
    """下载待认领收款Excel模板"""
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "待认领收款模板"

        # 设置表头
        headers = ['付款方式', '付款金额', '付款方', '收款主体', '商户订单号', '到账时间']
        ws.append(headers)

        # 设置表头样式
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # 添加示例数据
        ws.append(['微信', '1000.00', '张三', '北京', 'M202601120001', '2026-01-12'])

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='待认领收款模板.xlsx'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 导入待认领Excel数据
@app.route('/api/unclaimed/import', methods=['POST'])
def import_unclaimed_excel():
    """导入待认领收款Excel数据"""
    connection = None
    cursor = None
    try:
        # 检查文件
        if 'file' not in request.files:
            return jsonify({'error': '未上传文件'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': '文件名为空'}), 400

        # 检查文件格式
        if not file.filename.endswith(('.xls', '.xlsx')):
            return jsonify({'error': '仅支持.xls和.xlsx格式'}), 400

        # 读取Excel文件
        wb = load_workbook(file)
        ws = wb.active

        # 付款方式映射
        payment_method_map = {
            '微信': 0,
            '支付宝': 1,
            '优利支付': 2,
            '零零购支付': 3,
            '对公转账': 9
        }

        # 收款主体映射
        payee_entity_map = {
            '北京': 0,
            '西安': 1
        }

        connection = get_db_connection()
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # 跳过表头，从第二行开始读取
        success_count = 0
        matched_count = 0
        error_rows = []

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # 跳过空行
                continue

            try:
                payment_method_str = str(row[0]).strip() if row[0] else ''
                payment_amount_str = str(row[1]).strip() if row[1] else ''
                payer = str(row[2]).strip() if row[2] else ''
                payee_entity_str = str(row[3]).strip() if row[3] else ''
                merchant_order = str(row[4]).strip() if row[4] and str(row[4]).strip() != 'None' else None
                arrival_time_str = str(row[5]).strip() if row[5] else ''

                # 验证付款方式
                if payment_method_str not in payment_method_map:
                    error_rows.append(f'第{idx}行：付款方式不正确，仅支持：微信、支付宝、对公转账、零零购支付、优利支付')
                    continue

                payment_method = payment_method_map[payment_method_str]

                # 验证付款金额
                try:
                    payment_amount = float(payment_amount_str)
                    if payment_amount <= 0:
                        raise ValueError()
                    # 保留两位小数
                    payment_amount = round(payment_amount, 2)
                except:
                    error_rows.append(f'第{idx}行：付款金额格式不正确，必须为正数')
                    continue

                # 验证收款主体
                if payee_entity_str not in payee_entity_map:
                    error_rows.append(f'第{idx}行：收款主体不正确，仅支持：北京、西安')
                    continue

                payee_entity = payee_entity_map[payee_entity_str]

                # 验证到账时间格式
                try:
                    # 尝试解析日期
                    if isinstance(row[5], datetime):
                        arrival_time = row[5]
                    else:
                        # 支持YYYY-MM-DD格式
                        if not re.match(r'^\d{4}-\d{2}-\d{2}$', arrival_time_str):
                            raise ValueError()
                        arrival_time = datetime.strptime(arrival_time_str, '%Y-%m-%d')
                except:
                    error_rows.append(f'第{idx}行：到账时间格式不正确，必须为YYYY-MM-DD格式')
                    continue

                # 自动匹配逻辑：查找已收款中符合条件的记录
                matched = False
                if merchant_order:  # 只有商户订单号不为空时才进行匹配
                    cursor.execute("""
                        SELECT id FROM payment_collection
                        WHERE payment_scenario = 1
                        AND status = 10
                        AND payment_method = %s
                        AND payment_amount = %s
                        AND merchant_order = %s
                        AND payee_entity = %s
                        LIMIT 1
                    """, (payment_method, payment_amount, merchant_order, payee_entity))

                    matched_payment = cursor.fetchone()
                    if matched_payment:
                        # 更新已收款状态为已支付，同时更新到账时间
                        cursor.execute("""
                            UPDATE payment_collection
                            SET status = 20, arrival_time = %s
                            WHERE id = %s
                        """, (arrival_time, matched_payment['id']))
                        matched = True
                        matched_count += 1

                # 如果没有匹配到，则插入到待认领表
                if not matched:
                    cursor.execute("""
                        INSERT INTO unclaimed (payment_method, payment_amount, payer, payee_entity, merchant_order, arrival_time, status)
                        VALUES (%s, %s, %s, %s, %s, %s, 0)
                    """, (payment_method, payment_amount, payer, payee_entity, merchant_order, arrival_time))

                success_count += 1

            except Exception as e:
                error_rows.append(f'第{idx}行：{str(e)}')
                continue

        connection.commit()

        message = f'导入完成，成功{success_count}条'
        if matched_count > 0:
            message += f'，其中{matched_count}条自动匹配到已收款'

        if error_rows:
            return jsonify({
                'message': message,
                'success_count': success_count,
                'matched_count': matched_count,
                'errors': error_rows
            }), 200 if success_count > 0 else 400
        else:
            return jsonify({
                'message': message,
                'success_count': success_count,
                'matched_count': matched_count
            }), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'导入失败：{str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

# 临时API：创建unclaimed表
@app.route('/api/create-unclaimed-table', methods=['POST'])
def create_unclaimed_table():
    """创建unclaimed表"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        cursor = connection.cursor()

        sql = """
        CREATE TABLE IF NOT EXISTS `unclaimed` (
          `id` INT AUTO_INCREMENT PRIMARY KEY COMMENT '待认领ID',
          `payment_method` TINYINT NOT NULL DEFAULT 0 COMMENT '付款方式：0-微信，1-支付宝，2-优利支付，3-零零购支付，9-对公转账',
          `payment_amount` DECIMAL(10,2) NOT NULL DEFAULT 0.00 COMMENT '付款金额',
          `payer` VARCHAR(100) DEFAULT NULL COMMENT '付款方',
          `payee_entity` TINYINT NOT NULL DEFAULT 0 COMMENT '收款主体：0-北京，1-西安',
          `arrival_time` DATETIME DEFAULT NULL COMMENT '到账时间',
          `claimer` INT DEFAULT NULL COMMENT '认领人ID',
          `status` TINYINT NOT NULL DEFAULT 0 COMMENT '状态：0-待认领，1-已认领',
          `create_time` DATETIME DEFAULT CURRENT_TIMESTAMP COMMENT '创建时间',
          `update_time` DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP COMMENT '更新时间',
          INDEX `idx_status` (`status`),
          INDEX `idx_payment_method` (`payment_method`),
          INDEX `idx_arrival_time` (`arrival_time`)
        ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COMMENT='待认领收款表'
        """

        cursor.execute(sql)
        connection.commit()

        return jsonify({'message': 'unclaimed表创建成功'}), 200

    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5001)
